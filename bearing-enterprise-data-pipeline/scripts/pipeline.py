#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bearing-enterprise-data-pipeline: 企业产业标签全流程调度整合

严格按照 crawl → clean → tag 三步流水线自动执行，
统一归档原始数据、结构化数据、全维度标签，输出标准化JSON。

使用方式:
  python pipeline.py "企业名称"
  python pipeline.py "企业1,企业2,企业3"          # 批量
  python pipeline.py --file enterprises.txt        # 从文件读取批量
"""

import asyncio
import json
import os
import re
import sys
import time
import importlib.util
from datetime import datetime
from typing import Any, Dict, List, Optional

sys.stdout.reconfigure(encoding='utf-8')


# ============================================================
# 标准化进度展示器
# ============================================================
class ProgressReporter:
    """标准化进度展示器 - 对话框友好，只输出关键里程碑"""

    # 里程碑节点：只有这些进度值才会输出到控制台
    # 格式: progress% -> 是否输出
    MILESTONES = {
        0: True,    # 启动
        5: True,    # 采集开始
        25: False,  # 工商（子步骤，不输出）
        42: False,  # 招投标（子步骤）
        50: False,  # 专利（子步骤）
        60: True,   # 采集完成
        62: True,   # 清洗开始
        74: False,  # 资质（子步骤）
        80: True,   # 清洗完成
        82: True,   # 打标开始
        90: False,  # 能力标签（子步骤）
        95: True,   # 打标完成
        97: True,   # 归档
        100: True,  # 全部完成
    }

    # 阶段图标映射（避免emoji兼容性问题，用文字标记）
    STAGE_ICONS = {
        "调度": "[>]",
        "模块1采集": "[1]",
        "采集": "[1]",
        "模块2清洗": "[2]",
        "清洗": "[2]",
        "模块3打标": "[3]",
        "打标": "[3]",
    }

    def __init__(self, enterprise_name: str = ""):
        self.enterprise_name = enterprise_name
        self.current_progress = 0
        self.start_time = time.time()
        self._last_stage = ""
        self._stage_start_times = {}  # 各阶段开始时间

    def _format_time(self, seconds: float) -> str:
        """格式化耗时"""
        if seconds < 60:
            return f"{seconds:.0f}s"
        minutes = int(seconds // 60)
        secs = seconds % 60
        return f"{minutes}m{secs:.0f}s"

    def _get_icon(self, stage: str) -> str:
        """获取阶段图标"""
        for key, icon in self.STAGE_ICONS.items():
            if key in stage:
                return icon
        return "[*]"

    def report(self, progress: int, stage: str, step: str, detail: str = ""):
        """
        进度报告 - 只在里程碑节点输出

        Args:
            progress: 0-100 进度百分比
            stage: 模块阶段
            step: 具体步骤
            detail: 详细说明（可选）
        """
        self.current_progress = progress
        elapsed = time.time() - self.start_time

        # 记录阶段开始时间
        if stage != self._last_stage:
            self._stage_start_times[stage] = time.time()
            self._last_stage = stage

        # 只在里程碑节点输出
        is_milestone = progress in self.MILESTONES and self.MILESTONES[progress]
        # 阶段完成类报告也输出（detail含"完成"或"状态="）
        is_completion = detail and ("完成" in detail or "状态=" in detail or "置信度" in detail)
        # 失败报告必须输出
        is_failure = "失败" in step or "异常" in step

        if not (is_milestone or is_completion or is_failure):
            return

        icon = self._get_icon(stage)
        time_str = self._format_time(elapsed)
        line = f"  {icon} {progress:3d}% | {step}"

        if detail:
            line += f" -- {detail}"

        line += f"  ({time_str})"
        print(line)

    def report_pipeline_start(self, enterprise_name: str):
        """流水线启动提示"""
        self.enterprise_name = enterprise_name
        self.start_time = time.time()
        print(f"\n  [>] 企业产业标签全流程调度")
        print(f"  [>] 企业: {enterprise_name}")
        print(f"  [>] 流水线: 采集 -> 清洗 -> 打标 -> 归档")
        print(f"  {'-'*50}")

    def report_pipeline_done(self, status: str, confidence: float, total_time: float, token_summary: dict = None):
        """流水线完成提示"""
        time_str = self._format_time(total_time)
        status_icon = "[OK]" if status == "success" else "[!!]" if status == "partial" else "[X]"

        print(f"  {'-'*50}")
        print(f"  {status_icon} 流水线执行完成")
        print(f"  [>] 状态: {status} | 置信度: {confidence} | 耗时: {time_str}")
        if token_summary:
            total = token_summary.get("total_tokens", 0)
            print(f"  [>] Token用量: {total:,} (输入: {token_summary.get('total_input_tokens', 0):,}, 输出: {token_summary.get('total_output_tokens', 0):,})")
            for stage, usage in token_summary.get("stages", {}).items():
                print(f"       - {stage}: {usage['total_tokens']:,} tokens (入{usage['input_tokens']:,} + 出{usage['output_tokens']:,})")

    def report_pipeline_failed(self, stage: str, error: str, total_time: float):
        """流水线失败提示"""
        time_str = self._format_time(total_time)

        print(f"  {'-'*50}")
        print(f"  [X] 流水线执行失败")
        print(f"  [X] 失败环节: {stage}")
        print(f"  [X] 失败原因: {error}")
        print(f"  [>] 已耗时: {time_str}")

    def report_batch_progress(self, current: int, total: int, name: str, status: str):
        """批量任务进度"""
        print(f"  [{current}/{total}] {name} - {status}")


# ============================================================
# 模块加载器 - 动态导入三个子模块的类
# ============================================================
class ModuleLoader:
    """动态加载 bearing-enterprise-data-* 系列模块"""

    # 默认模块搜索路径（与当前skill同级目录）
    DEFAULT_SKILL_ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

    @classmethod
    def _find_skill_dir(cls, skill_name: str) -> Optional[str]:
        """查找 skill 目录"""
        # 1. 同级目录
        path = os.path.join(cls.DEFAULT_SKILL_ROOT, skill_name)
        if os.path.isdir(path):
            return path
        # 2. 当前工作目录下的
        path = os.path.join(os.getcwd(), skill_name)
        if os.path.isdir(path):
            return path
        return None

    @classmethod
    def load_crawler(cls) -> type:
        """加载 EnterpriseDataCrawler 类"""
        skill_dir = cls._find_skill_dir("bearing-enterprise-data-crawl")
        if not skill_dir:
            raise FileNotFoundError("未找到 bearing-enterprise-data-crawl skill 目录")
        script_path = os.path.join(skill_dir, "scripts", "crawl_enterprise.py")
        spec = importlib.util.spec_from_file_location("crawl_enterprise", script_path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod.EnterpriseDataCrawler

    @classmethod
    def load_cleaner(cls) -> type:
        """加载 BearingDataCleaner 类"""
        skill_dir = cls._find_skill_dir("bearing-enterprise-data-clean")
        if not skill_dir:
            raise FileNotFoundError("未找到 bearing-enterprise-data-clean skill 目录")
        script_path = os.path.join(skill_dir, "scripts", "clean_enterprise_data.py")
        spec = importlib.util.spec_from_file_location("clean_enterprise_data", script_path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod.BearingDataCleaner

    @classmethod
    def load_tagger(cls) -> type:
        """加载 EnterpriseTagger 类"""
        skill_dir = cls._find_skill_dir("bearing-enterprise-data-tag")
        if not skill_dir:
            raise FileNotFoundError("未找到 bearing-enterprise-data-tag skill 目录")
        script_path = os.path.join(skill_dir, "scripts", "tag_enterprise.py")
        spec = importlib.util.spec_from_file_location("tag_enterprise_data", script_path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod.EnterpriseTagger


# ============================================================
# 标签二次加工 - 从模块3标签体系提取扁平化标签
# ============================================================
class TagExtractor:
    """从模块3的三维标签体系中提取扁平化业务标签"""

    # 产品类型 → 高端标签映射
    HIGH_END_PRODUCTS = {
        "P4级精密轴承", "P2级超精密轴承", "陶瓷轴承", "混合陶瓷轴承",
        "航空轴承", "风电主轴轴承", "高铁轴承", "转盘轴承",
        "交叉滚子轴承", "薄壁轴承", "绝缘轴承",
        "高温轴承", "高速轴承", "真空轴承", "耐腐蚀轴承",
    }

    # 工艺能力关键词
    CRAFT_KEYWORDS = {
        "锻造": ["锻造技术", "自由锻", "模锻", "温锻", "热锻"],
        "车削": ["车削技术", "数控车削", "精密车削", "硬车削"],
        "磨削": ["磨削技术", "精密磨削", "内圆磨削", "外圆磨削", "无心磨削"],
        "超精加工": ["超精加工技术", "超精研", "抛光", "研磨"],
        "热处理": ["热处理技术", "淬火", "回火", "渗碳", "氮化", "感应淬火"],
        "装配": ["装配技术", "自动装配", "选配", "合套"],
    }

    @classmethod
    def extract_flat_tags(cls, tag_system: dict, structured_data: dict) -> dict:
        """从三维标签体系提取扁平化标签"""
        product_tags = tag_system.get("产品标签", {})
        service_tags = tag_system.get("服务标签", {})
        ability_tags = tag_system.get("能力标签", {})

        result = {}

        # ---- 产品维度 ----
        core_products = [t["tag"] for t in product_tags.get("核心产品类型", []) if t.get("confidence", 0) > 0]
        product_specs = [t["tag"] for t in product_tags.get("产品规格标签", []) if t.get("confidence", 0) > 0]
        app_scenes = [t["tag"] for t in product_tags.get("产品应用场景", []) if t.get("confidence", 0) > 0]
        compliance = [t["tag"] for t in product_tags.get("产品合规标签", []) if t.get("confidence", 0) > 0]

        main_products = cls._extract_main_products(structured_data)
        result["主营产品标签"] = main_products if main_products else core_products[:5]
        result["核心产品标签"] = core_products[:8] if core_products else ["未明确"]

        high_end = [p for p in core_products if p in cls.HIGH_END_PRODUCTS]
        for spec in product_specs:
            for kw in ["P4", "P2", "陶瓷", "混合", "航空级",
                       "风电级", "高铁级", "精密", "超精密", "高速", "真空"]:
                if kw in spec and spec not in high_end:
                    high_end.append(spec)
                    break
        result["高端产品标签"] = high_end if high_end else []

        keywords = list(set(core_products + compliance[:3]))
        result["产品关键词"] = keywords[:10] if keywords else ["未明确"]

        desc_parts = []
        if core_products:
            desc_parts.append(f"核心产品：{'、'.join(core_products[:5])}")
        if product_specs:
            desc_parts.append(f"产品规格：{'、'.join(product_specs[:5])}")
        if compliance:
            desc_parts.append(f"合规认证：{'、'.join(compliance[:3])}")
        result["产品结构描述"] = "；".join(desc_parts) if desc_parts else "产品信息未明确"

        # ---- 工艺维度 ----
        craft_tags = cls._extract_craft_tags(structured_data, ability_tags)
        result["工艺能力标签"] = craft_tags["craft_tags"]
        result["核心工艺标签"] = craft_tags["core_craft"]
        result["制造能力标签"] = craft_tags["manufacture_tags"]
        result["特种工艺标签"] = craft_tags["special_craft"]
        result["工艺能力描述"] = craft_tags["craft_desc"]

        # ---- 应用领域维度 ----
        result["应用领域标签"] = app_scenes if app_scenes else ["未明确"]
        result["核心应用领域"] = app_scenes[:3] if app_scenes else ["未明确"]
        downstream = cls._infer_downstream(structured_data, app_scenes)
        result["下游行业标签"] = downstream

        app_desc_parts = []
        if app_scenes:
            app_desc_parts.append(f"主要应用于{'、'.join(app_scenes[:5])}领域")
        coop = structured_data.get("cooperative_enterprise", "")
        if coop and coop not in ("未提取到", "未明确"):
            app_desc_parts.append(f"服务客户包括{coop}")
        result["应用领域描述"] = "；".join(app_desc_parts) if app_desc_parts else "应用领域未明确"

        # ---- 客户供应链维度 ----
        coop_modes = [t["tag"] for t in service_tags.get("合作模式标签", []) if t.get("confidence", 0) > 0]
        coverage = [t["tag"] for t in service_tags.get("服务覆盖范围", []) if t.get("confidence", 0) > 0]

        supply_chain_tags = cls._extract_supply_chain_tags(structured_data, coop_modes, app_scenes)
        result["客户供应链标签"] = supply_chain_tags
        result["客户类型标签"] = cls._infer_customer_types(structured_data, app_scenes)
        result["供应链角色标签"] = cls._infer_supply_chain_role(structured_data, coop_modes)
        result["供应链层级标签"] = cls._infer_supply_chain_tier(structured_data, coop_modes)

        sc_desc_parts = []
        coop_ent = structured_data.get("cooperative_enterprise", "")
        if coop_ent and coop_ent not in ("未提取到", "未明确"):
            sc_desc_parts.append(f"合作企业：{coop_ent}")
        if coop_modes:
            sc_desc_parts.append(f"合作模式：{'、'.join(coop_modes)}")
        if coverage:
            sc_desc_parts.append(f"覆盖范围：{'、'.join(coverage)}")
        result["客户供应链描述"] = "；".join(sc_desc_parts) if sc_desc_parts else "供应链信息未明确"

        return result

    @classmethod
    def _extract_main_products(cls, data: dict) -> List[str]:
        main_biz = data.get("main_business", "")
        core_products = data.get("core_products", "")
        biz_scope = data.get("business_scope", "")
        products = set()
        for text in [main_biz, core_products, biz_scope]:
            if not text or text in ("未提取到", "未明确"):
                continue
            for item in re.split(r'[,，、；;/\s]+', text):
                item = item.strip()
                if item and len(item) <= 15 and item not in (
                    "研发", "生产", "销售", "制造", "设计", "服务",
                    "加工", "开发", "咨询", "安装", "维修", "检测",
                    "及", "与", "及销售", "的生产", "的研发", "的制造"
                ):
                    products.add(item)
        return list(products)[:8]

    @classmethod
    def _extract_craft_tags(cls, data: dict, ability_tags: dict) -> dict:
        main_biz = data.get("main_business", "")
        core_products = data.get("core_products", "")
        biz_scope = data.get("business_scope", "")
        patent = data.get("patent_count", "")
        product_spec = data.get("product_spec", "")
        combined = f"{main_biz} {core_products} {biz_scope}"

        craft_tags, core_craft, manufacture_tags, special_craft = [], [], [], []

        for product_type, keywords in cls.CRAFT_KEYWORDS.items():
            if product_type in combined:
                for kw in keywords[1:]:
                    if kw in combined and kw not in craft_tags:
                        craft_tags.append(kw)

        tech_tags = [t["tag"] for t in ability_tags.get("技术能力标签", []) if t.get("confidence", 0) > 0]
        prod_tags = [t["tag"] for t in ability_tags.get("生产能力标签", []) if t.get("confidence", 0) > 0]

        for tag in tech_tags:
            if "技术" in tag and tag not in craft_tags:
                craft_tags.append(tag)
            if tag in ("高新技术企业", "专利密集型") and tag not in manufacture_tags:
                manufacture_tags.append(tag)
        for tag in prod_tags:
            if tag not in manufacture_tags:
                manufacture_tags.append(tag)

        if patent and patent not in ("未提取到", "未明确"):
            if "轴承" in patent and "轴承设计技术" not in core_craft:
                core_craft.append("轴承设计技术")
            if "密封" in patent and "密封技术" not in core_craft:
                core_craft.append("密封技术")
            if "润滑" in patent and "润滑技术" not in core_craft:
                core_craft.append("润滑技术")

        if product_spec and product_spec not in ("未提取到", "未明确"):
            spec_craft = {
                "P4": "精密加工工艺", "P2": "超精密加工工艺",
                "陶瓷": "陶瓷加工工艺", "混合": "混合陶瓷工艺",
                "真空": "真空冶炼工艺", "渗碳": "渗碳热处理工艺",
                "氮化": "氮化热处理工艺", "感应": "感应淬火工艺",
            }
            for kw, label in spec_craft.items():
                if kw in product_spec and label not in special_craft:
                    special_craft.append(label)

        special_keywords = {
            "高速": "高速工艺", "超高速": "超高速工艺", "耐高温": "耐高温工艺",
            "耐腐蚀": "耐腐蚀工艺", "真空": "真空工艺", "洁净": "洁净工艺",
            "航空": "航空级工艺", "风电": "风电级工艺", "高铁": "高铁级工艺",
            "军工": "军工级工艺", "核电": "核电级工艺",
        }
        for kw, label in special_keywords.items():
            if kw in combined and label not in special_craft:
                special_craft.append(label)

        if not craft_tags:
            craft_tags = ["未明确"]
        if not core_craft:
            core_craft = craft_tags[:3]
        if not manufacture_tags:
            manufacture_tags = ["未明确"]

        desc_parts = []
        if core_craft:
            desc_parts.append(f"核心工艺：{'、'.join(core_craft[:3])}")
        if manufacture_tags:
            desc_parts.append(f"制造能力：{'、'.join(manufacture_tags[:3])}")
        if special_craft:
            desc_parts.append(f"特种工艺：{'、'.join(special_craft[:3])}")

        return {
            "craft_tags": craft_tags[:8], "core_craft": core_craft[:5],
            "manufacture_tags": manufacture_tags[:8], "special_craft": special_craft[:5],
            "craft_desc": "；".join(desc_parts) if desc_parts else "工艺能力未明确",
        }

    @classmethod
    def _infer_downstream(cls, data: dict, app_scenes: List[str]) -> List[str]:
        scene_to_industry = {
            "汽车": "汽车工业", "风电": "风力发电行业",
            "航空航天": "航空航天行业", "铁路": "轨道交通行业",
            "矿山机械": "矿山机械行业", "机床": "机床工具行业",
            "工业机器人": "机器人行业", "冶金设备": "冶金行业",
            "石油装备": "石油装备行业", "农业机械": "农业机械行业",
            "工程机械": "工程机械行业", "电梯": "电梯行业",
            "家用电器": "家电行业", "电机": "电机制造行业",
            "泵阀": "泵阀制造行业", "船舶": "船舶制造行业",
            "海洋装备": "海洋装备行业",
            "高端制造": "高端装备制造业", "轴承": "轴承制造行业",
            "电力设备": "电力设备行业",
            "汽车工业": "汽车工业", "轨道交通": "轨道交通行业",
            "风力发电": "风力发电行业",
        }
        downstream = {scene_to_industry[s] for s in app_scenes if s in scene_to_industry}
        return list(downstream) if downstream else ["未明确"]

    @classmethod
    def _extract_supply_chain_tags(cls, data: dict, coop_modes: List[str], app_scenes: List[str]) -> List[str]:
        tags = []
        coop_ent = data.get("cooperative_enterprise", "")
        bidding = data.get("bidding_projects", "")
        famous_list = ["SKF", "NSK", "NTN", "铁姆肯", "Timken", "舍弗勒", "Schaeffler",
                       "FAG", "KOYO", "人本集团", "万向钱潮", "洛阳LYC", "瓦轴ZWZ",
                       "哈尔滨轴承HRB", "天马轴承", "五洲新春", "一汽", "中车",
                       "金风科技", "中航工业"]

        if coop_ent and coop_ent not in ("未提取到", "未明确"):
            coop_list = [e.strip() for e in re.split(r'[,，、；;]+', coop_ent) if e.strip()]
            if any(e in famous_list for e in coop_list):
                tags.append("头部企业供应商")
            if len(coop_list) >= 3:
                tags.append("多客户供应")
        if bidding and bidding not in ("未提取到", "未明确") and "中标" in bidding:
            tags.append("中标供应商")
        if "供应链配套" in coop_modes:
            tags.append("供应链配套商")
        if "批量供货" in coop_modes:
            tags.append("批量供应商")
        return tags if tags else ["未明确"]

    @classmethod
    def _infer_customer_types(cls, data: dict, app_scenes: List[str]) -> List[str]:
        types = []
        coop_ent = data.get("cooperative_enterprise", "")
        famous = ["SKF", "NSK", "NTN", "铁姆肯", "Timken", "舍弗勒",
                  "人本集团", "万向钱潮", "一汽", "中车", "金风科技"]
        if coop_ent and coop_ent not in ("未提取到", "未明确"):
            coop_list = [e.strip() for e in re.split(r'[,，、；;]+', coop_ent) if e.strip()]
            if any(e in famous for e in coop_list):
                types.append("头部企业客户")
            if len(coop_list) >= 3:
                types.append("多类型客户")
            types.append("B2B企业客户")
        scene_types = {
            "汽车": "车企客户", "铁路": "轨道交通客户",
            "风电": "风电设备商客户", "航空航天": "航空装备商客户",
            "工程机械": "工程机械客户", "机床": "机床制造商客户",
        }
        for scene, label in scene_types.items():
            if scene in app_scenes:
                types.append(label)
        return list(set(types)) if types else ["未明确"]

    @classmethod
    def _infer_supply_chain_role(cls, data: dict, coop_modes: List[str]) -> List[str]:
        roles = []
        combined = f"{data.get('main_business', '')} {data.get('business_scope', '')}"
        if "研发" in combined and "生产" in combined:
            roles.append("研产一体化供应商")
        elif "生产" in combined or "制造" in combined:
            roles.append("制造商")
        elif "研发" in combined or "设计" in combined:
            roles.append("设计服务商")
        if "锻造" in combined and "车削" in combined:
            roles.append("锻件供应商")
        if "代工" in combined or "OEM" in combined:
            roles.append("代工厂")
        if "供应链配套" in coop_modes:
            roles.append("配套供应商")
        if "定制化服务" in coop_modes:
            roles.append("定制化供应商")
        return list(set(roles)) if roles else ["供应商"]

    @classmethod
    def _infer_supply_chain_tier(cls, data: dict, coop_modes: List[str]) -> List[str]:
        tiers = []
        famous = ["SKF", "NSK", "NTN", "铁姆肯", "Timken", "舍弗勒", "人本集团"]
        coop_ent = data.get("cooperative_enterprise", "")
        capital = data.get("registered_capital", "")
        investment = data.get("investment_projects", "")

        if coop_ent and coop_ent not in ("未提取到", "未明确"):
            coop_list = [e.strip() for e in re.split(r'[,，、；;]+', coop_ent) if e.strip()]
            if any(e in famous for e in coop_list):
                tiers.append("一级供应商")
        if capital and capital not in ("未提取到", "未明确"):
            cap_yi = re.search(r'(\d+(?:\.\d+)?)\s*亿', capital)
            cap_wan = re.search(r'(\d+(?:\.\d+)?)\s*万', capital)
            cap_value = float(cap_yi.group(1)) * 10000 if cap_yi else (float(cap_wan.group(1)) if cap_wan else 0)
            if cap_value >= 5000:
                tiers.append("核心供应商")
        if investment and investment not in ("未提取到", "未明确"):
            tiers.append("战略供应商")
        return list(set(tiers)) if tiers else ["常规供应商"]


# ============================================================
# 企业事实关系提取器
# ============================================================
class FactRelationshipExtractor:
    """企业事实关系提取器 - 从原始爬虫文本+结构化数据+标签中智能提取5类核心关系
    
    核心改进：不再仅依赖硬编码字典，而是从原始文本中用上下文感知的方式提取企业名和关系。
    策略：原始文本正则提取 > 结构化字段 > 字典辅助映射
    """

    # ---- 辅助字典（用于行业映射，不再作为唯一来源） ----
    
    # 知名企业 → 行业映射（客户推断辅助）
    ENTERPRISE_INDUSTRY = {
        # 汽车
        "一汽": "汽车工业", "一汽集团": "汽车工业", "一汽解放": "汽车工业",
        "东风": "汽车工业", "东风汽车": "汽车工业", "东风日产": "汽车工业",
        "上汽": "汽车工业", "上汽集团": "汽车工业", "上汽大众": "汽车工业",
        "长安": "汽车工业", "长安汽车": "汽车工业",
        "广汽": "汽车工业", "广汽集团": "汽车工业",
        "吉利": "汽车工业", "吉利汽车": "汽车工业",
        "比亚迪": "汽车工业", "奇瑞": "汽车工业", "长城汽车": "汽车工业",
        "蔚来": "汽车工业", "理想": "汽车工业", "小鹏": "汽车工业",
        "北汽": "汽车工业", "江淮": "汽车工业", "重汽": "汽车工业",
        "陕汽": "汽车工业", "宇通": "汽车工业",
        # 轨道交通
        "中车": "轨道交通行业", "中国中车": "轨道交通行业",
        "中车株机": "轨道交通行业", "中车四方": "轨道交通行业",
        "中车长客": "轨道交通行业", "中车大连": "轨道交通行业",
        # 风电
        "金风科技": "风力发电行业", "明阳智能": "风力发电行业",
        "远景能源": "风力发电行业", "运达风电": "风力发电行业",
        "东方电气": "电力设备行业",
        # 工程机械
        "徐工": "工程机械行业", "徐工集团": "工程机械行业",
        "三一": "工程机械行业", "三一重工": "工程机械行业",
        "中联": "工程机械行业", "中联重科": "工程机械行业",
        "柳工": "工程机械行业", "山推": "工程机械行业",
        "临工": "工程机械行业",
        # 电力
        "西门子": "电力设备行业", "ABB": "电力设备行业",
        "上海电气": "电力设备行业", "哈电集团": "电力设备行业",
        # 家电
        "格力": "家电行业", "美的": "家电行业", "海尔": "家电行业",
        "海信": "家电行业", "TCL": "家电行业",
        # 冶金
        "宝钢": "冶金行业", "宝武": "冶金行业", "鞍钢": "冶金行业",
        "首钢": "冶金行业", "河钢": "冶金行业", "沙钢": "冶金行业",
        # 航空航天
        "中国商飞": "航空航天行业", "中航": "航空航天行业",
        "中航工业": "航空航天行业", "航天科技": "航空航天行业",
        "航发": "航空航天行业",
        # 机床
        "大连机床": "机床工具行业", "沈阳机床": "机床工具行业",
        "秦川机床": "机床工具行业", "齐重数控": "机床工具行业",
        # 矿山
        "中信重工": "矿山机械行业", "北方重工": "矿山机械行业",
        # 轴承同行
        "SKF": "轴承制造行业", "NSK": "轴承制造行业", "NTN": "轴承制造行业",
        "铁姆肯": "轴承制造行业", "Timken": "轴承制造行业",
        "舍弗勒": "轴承制造行业", "Schaeffler": "轴承制造行业",
        "FAG": "轴承制造行业", "INA": "轴承制造行业",
        "KOYO": "轴承制造行业", "NMB": "轴承制造行业",
        "Nachi": "轴承制造行业", "不二越": "轴承制造行业",
        "人本集团": "轴承制造行业", "万向钱潮": "轴承制造行业",
        "洛阳LYC": "轴承制造行业", "瓦轴": "轴承制造行业", "瓦轴ZWZ": "轴承制造行业",
        "哈轴": "轴承制造行业", "哈尔滨轴承": "轴承制造行业",
        "天马轴承": "轴承制造行业", "五洲新春": "轴承制造行业",
        "襄阳轴承": "轴承制造行业", "南方轴承": "轴承制造行业",
        "龙溪股份": "轴承制造行业", "宝塔实业": "轴承制造行业",
        "苏州轴承": "轴承制造行业", "西北轴承": "轴承制造行业",
        "大连冶金轴承": "轴承制造行业",
    }

    # 应用场景 → 行业映射
    SCENE_TO_INDUSTRY = {
        "汽车": "汽车工业", "汽车工业": "汽车工业", "汽车电子": "汽车工业",
        "风电": "风力发电行业", "风力发电": "风力发电行业",
        "航空航天": "航空航天行业", "航空": "航空航天行业",
        "铁路": "轨道交通行业", "高铁": "轨道交通行业", "轨道交通": "轨道交通行业",
        "矿山机械": "矿山机械行业", "矿山": "矿山机械行业",
        "机床": "机床工具行业", "数控": "机床工具行业",
        "工业机器人": "机器人行业", "机器人": "机器人行业",
        "冶金设备": "冶金行业", "冶金": "冶金行业", "轧机": "冶金行业",
        "石油装备": "石油装备行业", "石油": "石油装备行业",
        "农业机械": "农业机械行业", "农机": "农业机械行业",
        "工程机械": "工程机械行业", "挖掘机": "工程机械行业",
        "电梯": "电梯行业", "家电": "家电行业", "家用电器": "家电行业",
        "电机": "电机制造行业", "泵阀": "泵阀制造行业",
        "船舶": "船舶制造行业", "海洋装备": "海洋装备行业",
        "高端制造": "高端装备制造业", "轴承": "轴承制造行业",
        "电力设备": "电力设备行业", "电力": "电力设备行业",
        "压缩机": "通用机械行业", "减速机": "通用机械行业",
        "纺织": "纺织机械行业", "印刷": "印刷机械行业",
        "包装": "包装机械行业", "食品机械": "食品机械行业",
    }

    # ---- 关系上下文关键词 ----
    
    # 出现这些词 + 企业名 → 客户关系
    CUSTOMER_CONTEXT = [
        "供货", "供应", "提供", "配套", "交付", "交付给",
        "中标", "成交", "服务", "为客户", "向.*?提供",
        "采购方", "需求方", "终端客户",
    ]
    # 出现这些词 + 企业名 → 供应商关系
    SUPPLIER_CONTEXT = [
        "采购自", "采购于", "供应商", "由.*?供应", "向.*?采购",
        "原材料供应商", "零部件供应商", "进口",
    ]
    # 出现这些词 + 企业名 → 合作关系
    COOPERATION_CONTEXT = [
        "合作", "战略合作", "联合", "携手", "协作",
        "共同开发", "联合研发", "技术合作",
    ]
    # 招投标上下文
    BIDDING_CONTEXT = [
        "中标", "成交", "招标", "竞标", "投标",
    ]

    # 企业名提取正则模式（优先级从高到低）
    COMPANY_PATTERNS = [
        # 带括号的公司名（最精确）
        r'([\u4e00-\u9fa5A-Za-z]+(?:集团|股份|科技|技术|工业|装备|机械|电子|电气|汽车|轴承|精密|传动|冶金|重工|动力|机电|五金|钢铁|新材料|新能源)(?:有限)?(?:责任)?(?:公司|厂))',
        # "XX公司" 简写
        r'([\u4e00-\u9fa5]{2,8}(?:有限公司|公司|集团))',
        # 英文品牌名（大写2+字母）
        r'\b([A-Z][A-Z0-9]{1,10})\b',
    ]

    # 排除词（不是企业名的常见误匹配）
    EXCLUDE_NAMES = {
        "公司", "有限公司", "集团", "有限责任公司", "股份有限公司",
        "未提取到", "未明确", "暂无", "无", "不详",
        "APP", "PDF", "HTML", "VIP", "API", "ISO", "CEO", "CFO",
        "PCB", "LED", "FAQ", "URL", "GPS", "IOT", "ICT",
        "ICP", "SUJ2", "GCr15",
    }

    @classmethod
    def extract(cls, enterprise_name: str, structured_data: dict,
                enterprise_tags: dict, flat_tags: dict,
                raw_crawl_data: dict = None) -> List[dict]:
        """提取全部5类事实关系
        
        Args:
            enterprise_name: 当前企业名
            structured_data: 清洗后的结构化数据
            enterprise_tags: 模块3标签体系
            flat_tags: 扁平化标签
            raw_crawl_data: 原始爬虫数据（含原始文本，关键！）
        """
        facts = []
        
        # 获取原始文本（核心数据源）
        raw_text = cls._get_raw_text(raw_crawl_data)
        
        # Step1: 从原始文本中提取企业名及其上下文关系
        enterprise_relations = cls._extract_enterprises_from_raw(enterprise_name, raw_text)
        
        # Step2: 根据上下文分类关系类型
        facts.extend(cls._classify_and_build_facts(enterprise_name, enterprise_relations, structured_data, flat_tags))
        
        # Step3: 从结构化字段补充（招投标、投资等）
        facts.extend(cls._extract_bidding_relations(enterprise_name, structured_data, raw_text))
        facts.extend(cls._extract_investment_relations(enterprise_name, structured_data, raw_text))
        
        # Step4: 下游行业关系
        facts.extend(cls._extract_downstream_relations(enterprise_name, structured_data, flat_tags, enterprise_relations))
        
        # 去重
        facts = cls._deduplicate_facts(facts)
        
        return facts

    @classmethod
    def _get_raw_text(cls, raw_crawl_data: dict) -> str:
        """从原始爬虫数据中提取文本"""
        if not raw_crawl_data:
            return ""
        
        # 尝试多种数据结构
        raw_content = raw_crawl_data.get("raw_content", {})
        if raw_content:
            # 优先 all_content
            text = raw_content.get("all_content", "")
            if text:
                return text
            # 否则拼接各来源
            parts = []
            for key, value in raw_content.items():
                if value and isinstance(value, str):
                    parts.append(value)
            return "\n".join(parts)
        
        # 直接是字符串
        if isinstance(raw_crawl_data, str):
            return raw_crawl_data
        
        return ""

    @classmethod
    def _extract_enterprises_from_raw(cls, enterprise_name: str, raw_text: str) -> List[dict]:
        """从原始文本中提取所有企业名及其上下文关系线索
        
        Returns:
            [{"name": "XX公司", "context": "供货", "industry": "汽车工业", "source_snippet": "..."}]
        """
        if not raw_text:
            return []
        
        results = []
        seen_names = set()
        
        # 策略1: 用正则提取企业名
        for pattern in cls.COMPANY_PATTERNS:
            for m in re.finditer(pattern, raw_text):
                name = m.group(1).strip()
                if not cls._is_valid_enterprise_name(name, enterprise_name):
                    continue
                if name in seen_names:
                    continue
                
                # 获取上下文（前后各50个字符）
                ctx_start = max(0, m.start() - 50)
                ctx_end = min(len(raw_text), m.end() + 50)
                context = raw_text[ctx_start:ctx_end]
                
                # 判断上下文关系类型
                rel_context = cls._detect_context(context, name)
                # 判断行业
                industry = cls._infer_industry(name, context)
                
                seen_names.add(name)
                results.append({
                    "name": name,
                    "context": rel_context,
                    "industry": industry,
                    "source_snippet": context[:100],
                })
        
        # 策略2: 从已知企业字典中匹配（覆盖缩写/简称）
        for known_name, industry in cls.ENTERPRISE_INDUSTRY.items():
            if known_name in raw_text and known_name not in seen_names:
                if known_name == enterprise_name or known_name in enterprise_name:
                    continue
                # 获取上下文
                idx = raw_text.find(known_name)
                ctx_start = max(0, idx - 50)
                ctx_end = min(len(raw_text), idx + len(known_name) + 50)
                context = raw_text[ctx_start:ctx_end]
                
                rel_context = cls._detect_context(context, known_name)
                seen_names.add(known_name)
                results.append({
                    "name": known_name,
                    "context": rel_context,
                    "industry": industry,
                    "source_snippet": context[:100],
                })
        
        return results

    @classmethod
    def _is_valid_enterprise_name(cls, name: str, enterprise_name: str) -> bool:
        """判断是否为有效的企业名（排除噪声）"""
        if not name or len(name) < 2:
            return False
        if name in cls.EXCLUDE_NAMES:
            return False
        # 排除自身
        if name == enterprise_name or name in enterprise_name or enterprise_name in name:
            return False
        # 排除纯数字
        if re.match(r'^[\d.]+$', name):
            return False
        # 排除太短的英文（大概率不是品牌）
        if re.match(r'^[A-Z]+$', name) and len(name) < 2:
            return False
        # 排除常见非企业名词
        noise_words = {"轴承", "产品", "技术", "质量", "管理", "服务", "制造", "生产",
                       "加工", "研发", "销售", "设计", "安装", "维修", "检测", "认证"}
        if name in noise_words:
            return False
        return True

    @classmethod
    def _detect_context(cls, context: str, name: str) -> str:
        """检测上下文中的关系类型关键词"""
        # 客户上下文
        for kw in cls.CUSTOMER_CONTEXT:
            if re.search(kw, context):
                return "客户"
        # 供应商上下文
        for kw in cls.SUPPLIER_CONTEXT:
            if re.search(kw, context):
                return "供应商"
        # 合作上下文
        for kw in cls.COOPERATION_CONTEXT:
            if re.search(kw, context):
                return "合作"
        # 招投标上下文
        for kw in cls.BIDDING_CONTEXT:
            if re.search(kw, context):
                return "招投标"
        return "合作"  # 默认合作关系

    @classmethod
    def _infer_industry(cls, name: str, context: str) -> str:
        """从企业名和上下文推断行业"""
        # 先从字典查
        if name in cls.ENTERPRISE_INDUSTRY:
            return cls.ENTERPRISE_INDUSTRY[name]
        # 部分匹配
        for known, industry in cls.ENTERPRISE_INDUSTRY.items():
            if known in name or name in known:
                return industry
        # 从上下文推断
        for scene, industry in cls.SCENE_TO_INDUSTRY.items():
            if scene in context:
                return industry
        # 从企业名关键词推断
        name_industry_map = {
            "汽车": "汽车工业", "轴承": "轴承制造行业", "重工": "工程机械行业",
            "冶金": "冶金行业", "电气": "电力设备行业", "电子": "电子信息行业",
            "机械": "机械制造行业", "精密": "精密制造行业", "科技": "科技行业",
            "传动": "传动设备行业", "装备": "装备制造行业", "钢铁": "冶金行业",
            "电机": "电机制造行业", "动力": "动力设备行业",
        }
        for kw, industry in name_industry_map.items():
            if kw in name:
                return industry
        return ""

    @classmethod
    def _classify_and_build_facts(cls, enterprise_name: str, enterprise_relations: List[dict],
                                   structured_data: dict, flat_tags: dict) -> List[dict]:
        """根据上下文分类关系并构建事实关系"""
        facts = []
        is_bearing_company = cls._is_bearing_enterprise(structured_data)
        
        for rel in enterprise_relations:
            name = rel["name"]
            context = rel["context"]
            industry = rel["industry"]
            snippet = rel["source_snippet"]
            
            # 判断关系类型
            if context == "客户" or (context == "招投标" and not cls._is_bearing_name(name)):
                rel_type = "客户"
                desc = f"{enterprise_name}为{name}提供产品/服务"
                confidence = 0.85 if context == "客户" else 0.80
                source = "原始文本-客户上下文" if context == "客户" else "招投标信息"
            elif context == "供应商" or (is_bearing_company and cls._is_bearing_name(name)):
                rel_type = "供应商"
                desc = f"{name}为{enterprise_name}提供轴承相关产品/技术"
                confidence = 0.82
                source = "原始文本-供应商上下文"
            elif context == "招投标":
                # 中标信息中的对方通常是客户
                rel_type = "客户"
                desc = f"{enterprise_name}与{name}存在招投标业务关系"
                confidence = 0.78
                source = "招投标信息"
            else:
                rel_type = "合作"
                desc = f"{enterprise_name}与{name}存在合作关系"
                confidence = 0.75
                source = "原始文本-合作上下文"
            
            # 金额提取
            amount = ""
            amt_match = re.search(r'([\d,.]+)\s*万?元', snippet)
            if amt_match:
                amount = amt_match.group(0)
            
            # 推断依据
            evidence = f"原始文本匹配；上下文:{context}"
            if industry:
                evidence += f"；行业:{industry}"
            
            facts.append({
                "主体企业": enterprise_name,
                "关系客体": name,
                "关系类型": rel_type,
                "关系描述": desc,
                "关联项目": "",
                "关联金额": amount,
                "行业领域": industry,
                "推断依据": evidence,
                "置信度": confidence,
                "数据来源": source,
            })
        
        return facts

    @classmethod
    def _is_bearing_enterprise(cls, structured_data: dict) -> bool:
        """判断主体企业是否为轴承企业"""
        industry = structured_data.get("industry_segment", "")
        industry_cat = structured_data.get("industry_category", "")
        return "轴承" in industry or "轴承" in industry_cat

    @classmethod
    def _is_bearing_name(cls, name: str) -> bool:
        """判断企业名是否为轴承相关企业"""
        bearing_keywords = ["轴承", "SKF", "NSK", "NTN", "FAG", "INA", "KOYO",
                           "NMB", "Nachi", "Timken", "铁姆肯", "舍弗勒",
                           "人本", "瓦轴", "哈轴", "LYC", "ZWZ", "HRB",
                           "滚子", "保持架", "钢球"]
        for kw in bearing_keywords:
            if kw in name:
                return True
        # 查字典
        if name in cls.ENTERPRISE_INDUSTRY:
            return cls.ENTERPRISE_INDUSTRY[name] == "轴承制造行业"
        return False

    @classmethod
    def _extract_bidding_relations(cls, enterprise_name: str, structured_data: dict,
                                   raw_text: str) -> List[dict]:
        """从招投标信息中提取关系"""
        facts = []
        bidding = structured_data.get("bidding_projects", "")
        
        # 合并结构化字段和原始文本中的招投标句子
        bidding_text = bidding if bidding and bidding not in ("未提取到", "未明确") else ""
        
        # 从原始文本中找招投标句子
        if raw_text:
            for sentence in re.split(r'[。；\n]', raw_text):
                sentence = sentence.strip()
                if not sentence or len(sentence) < 15:
                    continue
                has_bid = any(kw in sentence for kw in cls.BIDDING_CONTEXT)
                has_money = bool(re.search(r'[\d,.]+\s*万?元', sentence))
                has_company = bool(re.search(r'[\u4e00-\u9fa5]+(?:有限公司|公司|集团)', sentence))
                if has_bid and (has_money or has_company):
                    if sentence not in bidding_text:
                        bidding_text += "。" + sentence if bidding_text else sentence
        
        if not bidding_text:
            return facts
        
        # 提取招投标中的企业名
        for pattern in cls.COMPANY_PATTERNS:
            for m in re.finditer(pattern, bidding_text):
                name = m.group(1).strip()
                if not cls._is_valid_enterprise_name(name, enterprise_name):
                    continue
                
                # 获取上下文
                ctx_start = max(0, m.start() - 30)
                ctx_end = min(len(bidding_text), m.end() + 30)
                context = bidding_text[ctx_start:ctx_end]
                
                # 金额
                amount = ""
                amt_match = re.search(r'([\d,.]+)\s*万?元', context)
                if amt_match:
                    amount = amt_match.group(0)
                
                industry = cls._infer_industry(name, context)
                
                # 判断关系：如果在"中标"前出现 = 采购方 = 客户
                if "中标" in context:
                    pre_text = context[:context.find("中标")]
                    if name in pre_text:
                        rel_type = "客户"
                    else:
                        rel_type = "客户"
                else:
                    rel_type = "客户"
                
                facts.append({
                    "主体企业": enterprise_name,
                    "关系客体": name,
                    "关系类型": rel_type,
                    "关系描述": f"{enterprise_name}与{name}存在招投标业务关系",
                    "关联项目": context[:80],
                    "关联金额": amount,
                    "行业领域": industry,
                    "推断依据": f"招投标上下文:{context[:50]}",
                    "置信度": 0.78,
                    "数据来源": "招投标信息",
                })
        
        return facts

    @classmethod
    def _extract_investment_relations(cls, enterprise_name: str, structured_data: dict,
                                      raw_text: str) -> List[dict]:
        """从投资信息中提取关系"""
        facts = []
        investment = structured_data.get("investment_projects", "")
        
        # 合并原始文本中的投资句子
        invest_text = investment if investment and investment not in ("未提取到", "未明确") else ""
        
        if raw_text:
            for sentence in re.split(r'[。；\n]', raw_text):
                sentence = sentence.strip()
                if not sentence or len(sentence) < 15:
                    continue
                has_invest = any(kw in sentence for kw in ["投资", "建设", "新增产线", "扩建", "技改"])
                has_money = bool(re.search(r'[\d,.]+\s*[亿万]元', sentence))
                if has_invest and has_money and sentence not in invest_text:
                    invest_text += "。" + sentence if invest_text else sentence
        
        if not invest_text:
            return facts
        
        # 提取投资金额
        amount = ""
        amt_yi = re.search(r'(\d+(?:\.\d+)?)\s*亿', invest_text)
        amt_wan = re.search(r'(\d+(?:\.\d+)?)\s*万', invest_text)
        if amt_yi:
            amount = f"{amt_yi.group(1)}亿元"
        elif amt_wan:
            amount = f"{amt_wan.group(1)}万元"
        
        # 提取投资项目摘要
        # 取第一个有效句子
        for sentence in re.split(r'[。；\n]', invest_text):
            sentence = sentence.strip()
            if len(sentence) >= 15 and any(kw in sentence for kw in ["投资", "建设", "新增", "扩建", "技改"]):
                facts.append({
                    "主体企业": enterprise_name,
                    "关系客体": sentence[:60] if len(sentence) > 60 else sentence,
                    "关系类型": "投资",
                    "关系描述": f"{enterprise_name}投资建设: {sentence[:80]}",
                    "关联项目": sentence[:80],
                    "关联金额": amount,
                    "行业领域": structured_data.get("industry_segment", ""),
                    "推断依据": "投资项目/原始文本",
                    "置信度": 0.85,
                    "数据来源": "投资信息",
                })
                break  # 只取第一个最相关的
        
        return facts

    @classmethod
    def _extract_downstream_relations(cls, enterprise_name: str, structured_data: dict,
                                       flat_tags: dict, enterprise_relations: List[dict]) -> List[dict]:
        """提取下游行业关系"""
        facts = []
        downstream = flat_tags.get("下游行业标签", [])
        app_scenes = flat_tags.get("应用领域标签", [])
        
        # 从标签提取
        if isinstance(downstream, list):
            for industry in downstream:
                if industry == "未明确" or not industry:
                    continue
                evidence_parts = []
                for scene in (app_scenes if isinstance(app_scenes, list) else []):
                    if cls.SCENE_TO_INDUSTRY.get(scene) == industry:
                        evidence_parts.append(f"应用场景:{scene}")
                # 从原始文本提取的企业关系中补充推断依据
                for rel in enterprise_relations:
                    if rel.get("industry") == industry:
                        evidence_parts.append(f"关联企业:{rel['name']}")
                
                evidence = "；".join(evidence_parts) if evidence_parts else "应用场景/业务范围推断"
                
                facts.append({
                    "主体企业": enterprise_name,
                    "关系客体": industry,
                    "关系类型": "下游行业",
                    "关系描述": f"{enterprise_name}产品应用于{industry}",
                    "关联项目": "",
                    "关联金额": "",
                    "行业领域": industry,
                    "推断依据": evidence,
                    "置信度": 0.85,
                    "数据来源": "标签推断",
                })
        
        # 补充：从原始文本中提取的企业关系推断出更多下游行业
        existing_industries = {f.get("关系客体") for f in facts}
        for rel in enterprise_relations:
            industry = rel.get("industry", "")
            if industry and industry not in existing_industries and industry != "轴承制造行业":
                existing_industries.add(industry)
                facts.append({
                    "主体企业": enterprise_name,
                    "关系客体": industry,
                    "关系类型": "下游行业",
                    "关系描述": f"{enterprise_name}产品应用于{industry}",
                    "关联项目": "",
                    "关联金额": "",
                    "行业领域": industry,
                    "推断依据": f"关联企业:{rel['name']}推断",
                    "置信度": 0.75,
                    "数据来源": "原始文本推断",
                })
        
        return facts

    @classmethod
    def _deduplicate_facts(cls, facts: List[dict]) -> List[dict]:
        """去重：同一(主体, 客体, 类型)只保留置信度最高的"""
        seen = {}
        for f in facts:
            key = (f.get("主体企业", ""), f.get("关系客体", ""), f.get("关系类型", ""))
            if key not in seen or f.get("置信度", 0) > seen[key].get("置信度", 0):
                seen[key] = f
        return list(seen.values())


# ============================================================
# Excel 报告导出器
# ============================================================
class ExcelExporter:
    """将标签+事实关系导出为Excel文件（3个Sheet）"""

    @staticmethod
    def export(enterprise_name: str, structured_data: dict,
               enterprise_tags: dict, flat_tags: dict,
               fact_relationships: List[dict], output_dir: str,
               structured_facts: dict = None) -> str:
        """导出Excel报告，返回文件路径"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            print("  [!] openpyxl 未安装，跳过Excel导出。请运行: pip install openpyxl")
            return ""

        wb = Workbook()

        # Sheet1: 企业标签总览
        ws1 = wb.active
        ws1.title = "企业标签总览"
        ExcelExporter._write_tag_overview(ws1, enterprise_name, structured_data, enterprise_tags, flat_tags)

        # Sheet2: 企业事实关系
        ws2 = wb.create_sheet("企业事实关系")
        ExcelExporter._write_fact_relationships(ws2, fact_relationships)

        # Sheet3: 事实商业关系明细（新增）
        if structured_facts and structured_facts.get("records"):
            ws3 = wb.create_sheet("事实商业关系明细")
            ExcelExporter._write_fact_details(ws3, structured_facts)

        # Sheet4: 元数据说明
        ws4 = wb.create_sheet("元数据说明")
        ExcelExporter._write_metadata(ws4)

        # 保存文件
        os.makedirs(output_dir, exist_ok=True)
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', enterprise_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{safe_name}_{timestamp}_report.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb.save(filepath)
        return filepath

    @staticmethod
    def _style_header(ws, headers: list, col_widths: list = None):
        """应用表头样式"""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        if col_widths:
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    @staticmethod
    def _write_tag_overview(ws, enterprise_name: str, structured_data: dict,
                            enterprise_tags: dict, flat_tags: dict):
        """写入企业标签总览Sheet"""
        headers = [
            "企业名称", "简称", "行业细分", "主营产品", "核心产品",
            "高端产品", "核心工艺", "特种工艺", "应用领域", "下游行业",
            "供应链角色", "供应链层级", "客户类型", "合规认证",
            "制造能力", "产品关键词", "标签置信度", "数据来源"
        ]
        col_widths = [25, 15, 20, 30, 30, 25, 25, 20, 25, 25, 20, 15, 20, 20, 20, 25, 12, 20]

        ExcelExporter._style_header(ws, headers, col_widths)

        from openpyxl.styles import Alignment, Border, Side

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alignment = Alignment(vertical='center', wrap_text=True)

        def join_list(lst):
            if isinstance(lst, list):
                return "、".join(str(x) for x in lst if x and x != "未明确")
            return str(lst) if lst else ""

        # 提取合规认证标签
        compliance_tags = []
        product_tags = enterprise_tags.get("产品标签", {})
        for t in product_tags.get("产品合规标签", []):
            if t.get("confidence", 0) > 0:
                compliance_tags.append(t["tag"])

        row_data = [
            enterprise_name,
            structured_data.get("enterprise_short_name", ""),
            structured_data.get("industry_segment", ""),
            join_list(flat_tags.get("主营产品标签", [])),
            join_list(flat_tags.get("核心产品标签", [])),
            join_list(flat_tags.get("高端产品标签", [])),
            join_list(flat_tags.get("核心工艺标签", [])),
            join_list(flat_tags.get("特种工艺标签", [])),
            join_list(flat_tags.get("应用领域标签", [])),
            join_list(flat_tags.get("下游行业标签", [])),
            join_list(flat_tags.get("供应链角色标签", [])),
            join_list(flat_tags.get("供应链层级标签", [])),
            join_list(flat_tags.get("客户类型标签", [])),
            "、".join(compliance_tags) if compliance_tags else "",
            join_list(flat_tags.get("制造能力标签", [])),
            join_list(flat_tags.get("产品关键词", [])),
            structured_data.get("confidence", ""),
            structured_data.get("data_source", ""),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.alignment = alignment
            cell.border = thin_border

        # 冻结首行
        ws.freeze_panes = "A2"

    @staticmethod
    def _write_fact_relationships(ws, fact_relationships: List[dict]):
        """写入企业事实关系Sheet"""
        headers = [
            "主体企业", "关系客体", "关系类型", "关系描述",
            "关联项目", "关联金额", "行业领域", "推断依据", "置信度", "数据来源"
        ]
        col_widths = [25, 25, 10, 40, 30, 15, 20, 30, 10, 15]

        ExcelExporter._style_header(ws, headers, col_widths)

        from openpyxl.styles import Alignment, Border, Side, PatternFill

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alignment = Alignment(vertical='center', wrap_text=True)

        # 关系类型颜色区分
        type_colors = {
            "客户": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
            "供应商": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "投资": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "合作": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
            "下游行业": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        }

        for row_idx, fact in enumerate(fact_relationships, 2):
            values = [
                fact.get("主体企业", ""),
                fact.get("关系客体", ""),
                fact.get("关系类型", ""),
                fact.get("关系描述", ""),
                fact.get("关联项目", ""),
                fact.get("关联金额", ""),
                fact.get("行业领域", ""),
                fact.get("推断依据", ""),
                fact.get("置信度", ""),
                fact.get("数据来源", ""),
            ]
            rel_type = fact.get("关系类型", "")
            row_fill = type_colors.get(rel_type)

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = alignment
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

        # 冻结首行
        ws.freeze_panes = "A2"

    @staticmethod
    def _write_fact_details(ws, facts_result: dict):
        """写入事实商业关系明细Sheet"""
        from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

        headers = [
            "记录类型", "记录ID", "对方企业", "项目名称",
            "关系类型", "角色", "产品", "产品详情",
            "金额", "金额单位", "日期", "地区",
            "行业", "证据类型", "来源平台", "置信度"
        ]
        col_widths = [10, 10, 25, 30, 10, 8, 15, 15, 12, 8, 12, 10, 15, 12, 15, 8]

        ExcelExporter._style_header(ws, headers, col_widths)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alignment = Alignment(vertical='center', wrap_text=True)

        # 记录类型颜色区分
        type_colors = {
            "bidding": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),     # 蓝
            "customer": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),      # 绿
            "supplier": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),      # 黄
            "investment": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),    # 红
            "partnership": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),   # 紫
            "product_supply": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"), # 灰
        }

        # 类型中文映射
        type_names = {
            "bidding": "招投标", "customer": "客户关系",
            "supplier": "供应商关系", "investment": "投资项目",
            "partnership": "合作关系", "product_supply": "产品供货",
        }

        records = facts_result.get("records", [])
        # 按record_type分组，组内按置信度排序
        records.sort(key=lambda r: (r.get("record_type", ""), -r.get("confidence", 0)))

        for row_idx, record in enumerate(records, 2):
            rtype = record.get("record_type", "")

            # 统一获取对方企业名（不同类型字段名不同）
            counterparty = ""
            if rtype == "customer":
                counterparty = record.get("customer_name", "")
            elif rtype == "supplier":
                counterparty = record.get("supplier_name", "")
            elif rtype == "partnership":
                counterparty = record.get("partner_name", "")
            elif rtype == "bidding":
                counterparty = record.get("counterparty", "")
            else:
                counterparty = record.get("buyer", record.get("counterparty", ""))

            # 统一获取关系描述
            rel_type = type_names.get(rtype, rtype)
            role = record.get("role", record.get("relationship_nature", ""))
            products = record.get("products", record.get("products_supplied", ""))
            product_detail = ""
            if rtype == "bidding":
                product_detail = record.get("product_detail", "")
            elif rtype == "product_supply":
                product_detail = record.get("product_detail", record.get("spec", ""))

            amount = record.get("amount", "")
            amount_unit = record.get("amount_unit", "")
            date = record.get("bid_date", record.get("start_date", ""))
            region = record.get("region", "")
            industry = record.get("counterparty_industry",
                                  record.get("customer_industry",
                                             record.get("partner_industry", "")))
            evidence = record.get("evidence_type", "原始文本")
            platform = record.get("source_platform", "")
            confidence = record.get("confidence", "")

            values = [
                rel_type, record.get("record_id", ""),
                counterparty, record.get("project_name", ""),
                rel_type, role, products, product_detail,
                amount, amount_unit, date, region,
                industry, evidence, platform, confidence,
            ]

            row_fill = type_colors.get(rtype)
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = alignment
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

        ws.freeze_panes = "A2"

        # 添加汇总行
        summary = facts_result.get("summary", {})
        summary_row = len(records) + 3
        ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=f"共{summary.get('total_records', 0)}条记录")

        type_summary_parts = []
        for rtype, count in summary.items():
            if rtype.endswith("_records") and count > 0:
                type_name = type_names.get(rtype.replace("_records", ""), rtype)
                type_summary_parts.append(f"{type_name}:{count}")
        if type_summary_parts:
            ws.cell(row=summary_row + 1, column=1, value="分类统计")
            ws.cell(row=summary_row + 1, column=2, value="、".join(type_summary_parts))

    @staticmethod
    def _write_metadata(ws):
        """写入元数据说明Sheet"""
        headers = ["字段名", "所属Sheet", "类型", "说明"]
        col_widths = [20, 15, 10, 50]

        ExcelExporter._style_header(ws, headers, col_widths)

        from openpyxl.styles import Alignment, Border, Side

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alignment = Alignment(vertical='center', wrap_text=True)

        metadata = [
            # 企业标签总览字段
            ("企业名称", "企业标签总览", "文本", "企业全称"),
            ("简称", "企业标签总览", "文本", "企业简称（去掉有限公司后缀）"),
            ("行业细分", "企业标签总览", "文本", "轴承行业细分领域（如滚动轴承制造/滑动轴承制造等）"),
            ("主营产品", "企业标签总览", "文本", "企业主营产品列表"),
            ("核心产品", "企业标签总览", "文本", "企业核心产品类型标签"),
            ("高端产品", "企业标签总览", "文本", "高端/精密/特种产品标签"),
            ("核心工艺", "企业标签总览", "文本", "核心制造工艺标签"),
            ("特种工艺", "企业标签总览", "文本", "特殊/高端工艺标签"),
            ("应用领域", "企业标签总览", "文本", "产品应用领域标签"),
            ("下游行业", "企业标签总览", "文本", "下游行业标签"),
            ("供应链角色", "企业标签总览", "文本", "在供应链中的角色定位"),
            ("供应链层级", "企业标签总览", "文本", "供应商层级（一级/核心/常规等）"),
            ("客户类型", "企业标签总览", "文本", "客户类型标签"),
            ("合规认证", "企业标签总览", "文本", "产品合规认证信息"),
            ("制造能力", "企业标签总览", "文本", "制造能力标签"),
            ("产品关键词", "企业标签总览", "文本", "产品+合规关键词汇总"),
            ("标签置信度", "企业标签总览", "数值", "整体置信度(0-1)"),
            ("数据来源", "企业标签总览", "文本", "数据采集来源平台"),
            # 企业事实关系字段
            ("主体企业", "企业事实关系", "文本", "当前分析的企业名称"),
            ("关系客体", "企业事实关系", "文本", "关系对方（企业名/行业/项目）"),
            ("关系类型", "企业事实关系", "枚举", "关系类型：客户/供应商/投资/合作/下游行业"),
            ("关系描述", "企业事实关系", "文本", "关系的事实描述"),
            ("关联项目", "企业事实关系", "文本", "关联的具体项目名称"),
            ("关联金额", "企业事实关系", "文本", "关联的金额信息"),
            ("行业领域", "企业事实关系", "文本", "关系客体所属行业"),
            ("推断依据", "企业事实关系", "文本", "关系推断的数据依据"),
            ("置信度", "企业事实关系", "数值", "关系置信度(0-1)"),
            ("数据来源", "企业事实关系", "文本", "数据来源类型"),
        ]

        for row_idx, (name, sheet, dtype, desc) in enumerate(metadata, 2):
            values = [name, sheet, dtype, desc]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = alignment
                cell.border = thin_border


# ============================================================
# 流水线调度器
# ============================================================
class TokenTracker:
    """Token 用量统计器 - 估算各阶段数据等效 token 数"""

    # 中文约 1.5 字符/token，英文约 4 字符/token，取混合估算 2 字符/token
    CHARS_PER_TOKEN = 2.0

    @staticmethod
    def estimate_tokens(data: Any) -> int:
        """估算 JSON 数据的等效 token 数"""
        text = json.dumps(data, ensure_ascii=False) if not isinstance(data, str) else data
        return max(1, int(len(text) / TokenTracker.CHARS_PER_TOKEN))

    def __init__(self):
        self.stages = {}
        self.total_input_tokens = 0
        self.total_output_tokens = 0

    def record(self, stage: str, input_data: Any, output_data: Any):
        """记录单个阶段的 token 估算"""
        in_tokens = self.estimate_tokens(input_data)
        out_tokens = self.estimate_tokens(output_data)
        self.stages[stage] = {
            "input_tokens": in_tokens,
            "output_tokens": out_tokens,
            "total_tokens": in_tokens + out_tokens,
        }
        self.total_input_tokens += in_tokens
        self.total_output_tokens += out_tokens

    def summary(self) -> dict:
        """返回 token 用量汇总"""
        return {
            "stages": self.stages,
            "total_input_tokens": self.total_input_tokens,
            "total_output_tokens": self.total_output_tokens,
            "total_tokens": self.total_input_tokens + self.total_output_tokens,
        }


class PipelineOrchestrator:
    """企业产业标签全流程调度器"""

    def __init__(self, output_dir: str = None, llm_client=None):
        self.log_records = []
        self.module_root = ModuleLoader.DEFAULT_SKILL_ROOT
        self.progress = ProgressReporter()
        self.output_files = {}  # 记录各阶段输出文件路径
        self.token_tracker = TokenTracker()  # token 用量追踪
        self.llm_client = llm_client  # 可选的 LLM 客户端
        # 输出目录：优先使用指定目录，否则使用当前工作目录下的 output
        # 输出目录优先级: --output-dir 参数 > PROJECT_DIR 环境变量 > 当前工作目录
        if output_dir:
            self.output_dir = output_dir
        elif os.environ.get("PROJECT_DIR"):
            self.output_dir = os.path.join(os.environ.get("PROJECT_DIR"), "output")
        else:
            self.output_dir = os.path.join(os.getcwd(), "output")

    def _log(self, node: str, message: str, status: str = "info"):
        """记录日志（静默，不输出到控制台，进度由ProgressReporter负责）"""
        record = {
            "node": node,
            "message": message,
            "status": status,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        self.log_records.append(record)

    def _progress_callback(self, progress: int, stage: str, step: str, detail: str = ""):
        """统一进度回调 - 各模块通过此回调报告进度"""
        self.progress.report(progress, stage, step, detail)

    def run_single(self, enterprise_name: str) -> dict:
        """执行单企业全流程"""
        self.log_records = []
        start_time = time.time()
        self.progress = ProgressReporter(enterprise_name)
        self.token_tracker = TokenTracker()  # 每次执行重置 token 计数
        execute_nodes = []
        pipeline_status = "success"
        error_info = ""

        raw_crawl_data = {}
        structured_data = {}
        enterprise_tags = {}
        flat_tags = {}
        fact_relationships = []  # 企业事实关系
        tag_result = {}  # 模块3完整结果（含enterprise_info等）

        # 启动提示
        self.progress.report_pipeline_start(enterprise_name)
        self.progress.report(0, "调度", "启动流水线", f"目标企业: {enterprise_name}")

        # ==================== Step 1: 采集 ====================
        self.progress.report(2, "调度", "加载模块1", "初始化采集器")
        step1_start = time.time()

        try:
            CrawlerClass = ModuleLoader.load_crawler()
            crawler = CrawlerClass(max_pages=20, progress_callback=self._progress_callback, output_dir=self.output_dir)
            crawl_result_json = crawler.run(enterprise_name)
            raw_crawl_data = json.loads(crawl_result_json)
            self.token_tracker.record("采集", {"enterprise_name": enterprise_name}, raw_crawl_data)

            crawl_status = raw_crawl_data.get("crawl_status", "failed")
            step1_time = round(time.time() - step1_start, 2)

            if crawl_status in ("success", "partial"):
                execute_nodes.append("采集")
                self.progress.report(60, "采集", "采集完成",
                                     f"状态={crawl_status}, 置信度={raw_crawl_data.get('confidence', 0)}, 耗时{step1_time}s")
                self._log("采集", f"采集完成，状态={crawl_status}，耗时{step1_time}s", "success")
            else:
                pipeline_status = "failed"
                error_info = f"采集失败: crawl_status={crawl_status}"
                self._log("采集", f"采集失败，状态={crawl_status}，耗时{step1_time}s", "error")
                self.progress.report_pipeline_failed("采集", error_info, time.time() - start_time)
                return self._build_result(
                    enterprise_name, pipeline_status, execute_nodes,
                    raw_crawl_data, structured_data, enterprise_tags, flat_tags,
                    error_info, start_time
                )
        except Exception as e:
            pipeline_status = "failed"
            error_info = f"采集异常: {str(e)}"
            self._log("采集", f"采集异常: {str(e)}", "error")
            self.progress.report_pipeline_failed("采集", error_info, time.time() - start_time)
            return self._build_result(
                enterprise_name, pipeline_status, execute_nodes,
                raw_crawl_data, structured_data, enterprise_tags, flat_tags,
                error_info, start_time
            )

        # ==================== Step 2: 清洗 ====================
        step2_start = time.time()

        try:
            CleanerClass = ModuleLoader.load_cleaner()
            cleaner = CleanerClass(progress_callback=self._progress_callback, output_dir=self.output_dir, llm_client=self.llm_client)
            clean_result = cleaner.clean(raw_crawl_data)
            structured_data = clean_result
            self.token_tracker.record("清洗", raw_crawl_data, clean_result)

            clean_status = clean_result.get("clean_status", "failed")
            step2_time = round(time.time() - step2_start, 2)

            if clean_status in ("success", "partial"):
                execute_nodes.append("清洗")
                self.progress.report(80, "清洗", "清洗完成",
                                     f"状态={clean_status}, 置信度={clean_result.get('confidence', 0)}, 耗时{step2_time}s")
                self._log("清洗", f"清洗完成，状态={clean_status}，耗时{step2_time}s", "success")
            else:
                pipeline_status = "failed"
                error_info = f"清洗失败: clean_status={clean_status}"
                self._log("清洗", f"清洗失败，状态={clean_status}，耗时{step2_time}s", "error")
                self.progress.report_pipeline_failed("清洗", error_info, time.time() - start_time)
                return self._build_result(
                    enterprise_name, pipeline_status, execute_nodes,
                    raw_crawl_data, structured_data, enterprise_tags, flat_tags,
                    error_info, start_time
                )
        except Exception as e:
            pipeline_status = "failed"
            error_info = f"清洗异常: {str(e)}"
            self._log("清洗", f"清洗异常: {str(e)}", "error")
            self.progress.report_pipeline_failed("清洗", error_info, time.time() - start_time)
            return self._build_result(
                enterprise_name, pipeline_status, execute_nodes,
                raw_crawl_data, structured_data, enterprise_tags, flat_tags,
                error_info, start_time
            )

        # ==================== Step 3: 标签生成 ====================
        step3_start = time.time()

        try:
            TaggerClass = ModuleLoader.load_tagger()
            tagger = TaggerClass(progress_callback=self._progress_callback, output_dir=self.output_dir, llm_client=self.llm_client)
            tag_result = tagger.generate_tags(structured_data)
            self.token_tracker.record("打标", structured_data, tag_result)

            enterprise_tags = tag_result.get("tag_system", {})
            step3_time = round(time.time() - step3_start, 2)
            execute_nodes.append("打标")

            self.progress.report(94, "打标", "标签生成完成",
                                 f"置信度={tag_result.get('enterprise_info', {}).get('tag_confidence', 0)}, 耗时{step3_time}s")
            self._log("打标", f"标签生成完成，耗时{step3_time}s", "success")

            # 提取扁平化标签
            self.progress.report(94, "调度", "扁平化标签提取", "从三维标签体系提取业务标签")
            flat_tags = TagExtractor.extract_flat_tags(enterprise_tags, structured_data)
            self._log("调度", "扁平化标签提取完成", "success")

        except Exception as e:
            pipeline_status = "partial"
            error_info = f"打标异常: {str(e)}"
            self._log("打标", f"标签生成异常: {str(e)}", "error")

        # ==================== Step 4: 事实关系提取 ====================
        try:
            self.progress.report(95, "调度", "事实关系提取", "提取客户/供应商/投资/合作/下游行业关系")
            fact_relationships = FactRelationshipExtractor.extract(
                enterprise_name, structured_data, enterprise_tags, flat_tags,
                raw_crawl_data=raw_crawl_data
            )
            self._log("调度", f"事实关系提取完成，共{len(fact_relationships)}条关系", "success")
        except Exception as e:
            self._log("调度", f"事实关系提取异常: {str(e)}", "warning")

        # ==================== Step 5: 事实数据结构化提取（新增） ====================
        structured_facts = {"records": [], "summary": {"total_records": 0}}
        try:
            self.progress.report(95, "调度", "事实数据提取", "提取招投标/客户/供应商/投资结构化记录")
            # 动态导入 FactDataExtractor
            try:
                _project_root = ModuleLoader.DEFAULT_SKILL_ROOT
                if _project_root not in sys.path:
                    sys.path.insert(0, _project_root)
                from common.fact_extractor import FactDataExtractor
            except ImportError:
                # 回退：尝试从pipeline.py同级的common目录导入
                _pipeline_common = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "common")
                if os.path.isdir(_pipeline_common) and _pipeline_common not in sys.path:
                    sys.path.insert(0, os.path.dirname(_pipeline_common))
                from common.fact_extractor import FactDataExtractor

            fact_extractor = FactDataExtractor(output_dir=self.output_dir)
            structured_facts = fact_extractor.extract(
                enterprise_name, raw_crawl_data, structured_data
            )
            self._log("调度",
                      f"事实数据提取完成，共{structured_facts['summary']['total_records']}条记录",
                      "success")
        except Exception as e:
            import traceback as _tb
            self._log("调度", f"事实数据提取异常: {str(e)}\n{_tb.format_exc()}", "warning")

        # ==================== 归档 ====================
        self.progress.report(97, "调度", "数据归档", "保存各阶段文件到output目录")
        total_time = round(time.time() - start_time, 2)

        # 保存各阶段过程文件（统一时间戳，文件按阶段后缀区分）
        self.output_files = {}
        archive_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', enterprise_name)

        # 保存模块1采集原始数据
        if raw_crawl_data:
            crawl_path = self._save_stage_file(raw_crawl_data, safe_name, archive_timestamp, "crawl")
            self.output_files["crawl"] = crawl_path

        # 保存模块2清洗结构化数据
        if structured_data:
            clean_path = self._save_stage_file(structured_data, safe_name, archive_timestamp, "clean")
            self.output_files["clean"] = clean_path

        # 保存模块3标签数据
        if enterprise_tags:
            tag_data = {
                "enterprise_info": tag_result.get("enterprise_info", {}),
                "tag_system": enterprise_tags,
                "uncertain_tags": tag_result.get("uncertain_tags", []),
                "note": tag_result.get("note", "")
            }
            tag_path = self._save_stage_file(tag_data, safe_name, archive_timestamp, "tag")
            self.output_files["tag"] = tag_path

        # 保存事实数据（独立JSON文件）
        if structured_facts and structured_facts.get("records"):
            facts_path = self._save_stage_file(structured_facts, safe_name, archive_timestamp, "facts")
            self.output_files["facts"] = facts_path
            self._log("归档", f"事实数据已保存: {facts_path}", "success")

        # Excel 报告导出
        if flat_tags or fact_relationships:
            self.progress.report(96, "调度", "Excel报告导出", "导出标签总览+事实关系到Excel")
            try:
                excel_path = ExcelExporter.export(
                    enterprise_name, structured_data, enterprise_tags,
                    flat_tags, fact_relationships, self.output_dir,
                    structured_facts=structured_facts
                )
                if excel_path:
                    self.output_files["report"] = excel_path
                    self._log("归档", f"Excel报告已导出: {excel_path}", "success")
            except Exception as e:
                self._log("归档", f"Excel导出异常: {str(e)}", "warning")

        # 计算总体置信度
        confidences = []
        if raw_crawl_data.get("confidence"):
            confidences.append(raw_crawl_data["confidence"])
        if structured_data.get("confidence"):
            confidences.append(structured_data["confidence"])
        if enterprise_tags:
            for dim_tags in enterprise_tags.values():
                for sub_tags in dim_tags.values():
                    if isinstance(sub_tags, list):
                        for t in sub_tags:
                            if isinstance(t, dict) and t.get("confidence", 0) > 0:
                                confidences.append(t["confidence"])
        overall_confidence = round(sum(confidences) / len(confidences), 2) if confidences else 0.0

        self.progress.report(100, "调度", "全流程完成", f"状态={pipeline_status}, 置信度={overall_confidence}")
        token_summary = self.token_tracker.summary()
        self.progress.report_pipeline_done(pipeline_status, overall_confidence, total_time, token_summary)

        return self._build_result(
            enterprise_name, pipeline_status, execute_nodes,
            raw_crawl_data, structured_data, enterprise_tags, flat_tags,
            error_info, start_time, fact_relationships, structured_facts
        )

    def _save_stage_file(self, data: dict, safe_name: str, timestamp: str, stage: str) -> str:
        """保存单个阶段的过程文件到 output 目录

        Args:
            data: 阶段数据
            safe_name: 企业安全名称
            timestamp: 统一时间戳
            stage: 阶段标识（crawl/clean/tag/pipeline）

        Returns:
            保存的文件路径
        """
        output_dir = self.output_dir
        os.makedirs(output_dir, exist_ok=True)

        filename = f"{safe_name}_{timestamp}_{stage}.json"
        filepath = os.path.join(output_dir, filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return filepath



    def _build_result(
        self,
        enterprise_name: str,
        pipeline_status: str,
        execute_nodes: List[str],
        raw_crawl_data: dict,
        structured_data: dict,
        enterprise_tags: dict,
        flat_tags: dict,
        error_info: str,
        start_time: float,
        fact_relationships: List[dict] = None,
        structured_facts: dict = None
    ) -> dict:
        """构建最终输出结果"""

        if fact_relationships is None:
            fact_relationships = []
        if structured_facts is None:
            structured_facts = {"records": [], "summary": {"total_records": 0}}

        # 计算总体置信度
        confidences = []
        if raw_crawl_data.get("confidence"):
            confidences.append(raw_crawl_data["confidence"])
        if structured_data.get("confidence"):
            confidences.append(structured_data["confidence"])
        if enterprise_tags:
            for dim_tags in enterprise_tags.values():
                for sub_tags in dim_tags.values():
                    if isinstance(sub_tags, list):
                        for t in sub_tags:
                            if isinstance(t, dict) and t.get("confidence", 0) > 0:
                                confidences.append(t["confidence"])
        overall_confidence = round(sum(confidences) / len(confidences), 2) if confidences else 0.0

        # 生成日志摘要
        error_nodes = [r for r in self.log_records if r["status"] == "error"]
        warning_nodes = [r for r in self.log_records if r["status"] == "warning"]
        if error_nodes:
            log_summary = f"流程异常，{len(error_nodes)}个节点报错：{'；'.join(r['message'] for r in error_nodes[:3])}"
        elif warning_nodes:
            log_summary = f"流程完成但有警告，{len(warning_nodes)}项：{'；'.join(r['message'] for r in warning_nodes[:3])}"
        else:
            log_summary = "全流程节点执行正常，链路数据完整无缺失"

        result = {
            "pipeline_info": {
                "pipeline_status": pipeline_status,
                "execute_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "execute_node": execute_nodes,
                "overall_confidence": overall_confidence,
                "total_time_seconds": round(time.time() - start_time, 2),
                "output_files": self.output_files,
                "token_usage": self.token_tracker.summary(),
            },
            "raw_crawl_data": raw_crawl_data,
            "structured_data": structured_data,
            "enterprise_tags": enterprise_tags,
            "tags": flat_tags,
            "fact_relationships": fact_relationships,
            "structured_facts": structured_facts,
            "error_info": error_info,
            "log_record": log_summary,
            "log_detail": self.log_records,
        }

        return result

    def save_to_file(self, data: dict, enterprise_name: str) -> str:
        """将 pipeline 最终结果写入项目的 output 目录"""
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', enterprise_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self._save_stage_file(data, safe_name, timestamp, "pipeline")

    def run_batch(self, enterprise_names: List[str]) -> List[dict]:
        """批量执行全流程"""
        results = []
        total = len(enterprise_names)

        print(f"\n  [>] 批量调度开始，共 {total} 家企业")
        print(f"  {'-'*50}")

        for i, name in enumerate(enterprise_names, 1):
            self.progress.report_batch_progress(i, total, name, "开始执行")
            result = self.run_single(name)

            # 保存单企业结果
            output_path = self.save_to_file(result, name)
            result["output_path"] = output_path

            results.append(result)

            status = result['pipeline_info']['pipeline_status']
            self.progress.report_batch_progress(i, total, name, f"{status} -> {output_path}")

        # 批量汇总
        success_count = sum(1 for r in results if r["pipeline_info"]["pipeline_status"] in ("success", "partial"))
        total_tokens = sum(r["pipeline_info"].get("token_usage", {}).get("total_tokens", 0) for r in results)
        print(f"\n  [>] 批量调度完成: {success_count}/{total} 家企业成功")
        if total_tokens:
            print(f"  [>] 批量Token总用量: {total_tokens:,} tokens")

        return results


# ============================================================
# CLI 入口
# ============================================================
def main():
    import argparse
    parser = argparse.ArgumentParser(description="企业产业标签全流程调度")
    parser.add_argument("enterprise_name", nargs="?", help="企业名称（单个或逗号分隔多个）")
    parser.add_argument("--file", help="从文件读取企业名单")
    parser.add_argument("--output-dir", help="输出目录（默认为当前工作目录下的 output）", default=None)
    args = parser.parse_args()

    if not args.enterprise_name and not args.file:
        parser.print_help()
        print("\n示例:")
        print('  python pipeline.py "企业名称"')
        print('  python pipeline.py "企业1,企业2,企业3"')
        print('  python pipeline.py --file enterprises.txt')
        print('  python pipeline.py "企业名称" --output-dir /path/to/project/output')
        sys.exit(1)

    orchestrator = PipelineOrchestrator(output_dir=args.output_dir)

    if args.file:
        file_path = args.file
        if not os.path.exists(file_path):
            print(f"错误: 文件不存在 - {file_path}")
            sys.exit(1)
        with open(file_path, 'r', encoding='utf-8') as f:
            names = [line.strip() for line in f if line.strip() and not line.startswith('#')]
        results = orchestrator.run_batch(names)
    elif "," in args.enterprise_name:
        names = [n.strip() for n in args.enterprise_name.split(",") if n.strip()]
        results = orchestrator.run_batch(names)
    else:
        enterprise_name = args.enterprise_name
        result = orchestrator.run_single(enterprise_name)

        # 保存结果
        output_path = orchestrator.save_to_file(result, enterprise_name)

        # 输出标签摘要
        if result.get("tags"):
            tags = result["tags"]
            print(f"\n  [>] 标签摘要:")
            print(f"      主营产品: {', '.join(tags.get('主营产品标签', []))}")
            print(f"      核心工艺: {', '.join(tags.get('核心工艺标签', []))}")
            print(f"      应用领域: {', '.join(tags.get('应用领域标签', []))}")
            print(f"      供应链角色: {', '.join(tags.get('供应链角色标签', []))}")

        # 输出事实关系摘要
        if result.get("fact_relationships"):
            facts = result["fact_relationships"]
            # 按关系类型统计
            type_counts = {}
            for f in facts:
                rt = f.get("关系类型", "未知")
                type_counts[rt] = type_counts.get(rt, 0) + 1
            print(f"\n  [>] 事实关系摘要 (共{len(facts)}条):")
            for rt, cnt in type_counts.items():
                # 列出该类型的客体
                objects = [f["关系客体"] for f in facts if f.get("关系类型") == rt]
                obj_str = "、".join(objects[:5])
                if len(objects) > 5:
                    obj_str += f"等{len(objects)}个"
                print(f"      {rt}({cnt}条): {obj_str}")

        if result.get("error_info"):
            print(f"\n  [!!] 异常: {result['error_info']}")

        # Token 用量摘要
        token_usage = result.get("pipeline_info", {}).get("token_usage", {})
        if token_usage:
            total = token_usage.get("total_tokens", 0)
            print(f"\n  [>] Token用量估算: {total:,} tokens")
            print(f"      输入: {token_usage.get('total_input_tokens', 0):,} | 输出: {token_usage.get('total_output_tokens', 0):,}")

        print(f"\n  [>] 结果已保存: {output_path}")
        # Excel报告路径
        report_path = result.get("pipeline_info", {}).get("output_files", {}).get("report", "")
        if report_path:
            print(f"  [>] Excel报告: {report_path}")
        print(f"  [>] {result['log_record']}")


if __name__ == "__main__":
    main()

